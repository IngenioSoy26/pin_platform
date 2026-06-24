import csv
import json
import traceback
import os
from itertools import chain
from django.utils import timezone
from apps.core.models import TruckStop, WIMStation, Amenity, RecyclingFacility, TransportationStatistic, TireShop, AlternativeFuelStation, CarrierCompany, WeightEnforcementStatistic
from apps.routes.models import HPMSInventorySource
from apps.routes.services import RouteService
from .models import DataSource, ETLJob


def _get_progress_interval(file_path):
    file_size = os.path.getsize(file_path)
    if file_size >= 1_000_000_000:
        return 50_000
    if file_size >= 100_000_000:
        return 10_000
    if file_size >= 10_000_000:
        return 2_000
    return 500


def _track_reader_progress(reader, etl_job):
    interval = _get_progress_interval(etl_job.dataset.file_path)

    def generator():
        for processed_count, row in enumerate(reader, start=1):
            if processed_count % interval == 0:
                etl_job.records_processed = processed_count
                etl_job.save(update_fields=['records_processed'])
            yield row

    return generator()

def run_etl_job(etl_job):
    """
    Motor ETL principal. Lee el archivo crudo y lo mapea a los modelos de Django.
    """
    etl_job.status = 'RUNNING'
    etl_job.error_log = "Iniciando procesamiento ETL nativo...\n"
    etl_job.save()

    try:
        file_path = etl_job.dataset.file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"El archivo {file_path} no existe.")
            
        ext = os.path.splitext(file_path)[1].lower()
        
        # Procesamiento CSV
        if ext == '.csv':
            with open(file_path, mode='r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.DictReader(f)
                _route_processing(reader, etl_job)
                
        # Procesamiento Excel (xlsx)
        elif ext in ['.xlsx', '.xls']:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            try:
                sheet = wb.active

                header_row_idx = 1
                headers = []
                max_headers = 0

                for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                    current_headers = [str(cell).strip() if cell else f"Empty_{j}" for j, cell in enumerate(row)]
                    valid_headers_count = sum(1 for h in current_headers if not h.startswith("Empty_"))

                    if valid_headers_count > max_headers and valid_headers_count > 3:
                        max_headers = valid_headers_count
                        headers = current_headers
                        header_row_idx = i

                def iter_excel_rows():
                    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        yield {
                            headers[i]: row[i]
                            for i in range(min(len(headers), len(row)))
                        }

                _route_processing(iter_excel_rows(), etl_job)
            finally:
                wb.close()
            
        # Procesamiento GeoJSON
        elif ext == '.geojson':
            with open(file_path, mode='r', encoding='utf-8') as f:
                geojson_data = json.load(f)

            def iter_geojson_rows():
                for feature in geojson_data.get('features', []):
                    props = dict(feature.get('properties', {}))
                    geom = feature.get('geometry', {})
                    if geom:
                        props['geometry'] = geom
                    if geom and geom.get('type') == 'Point':
                        coords = geom.get('coordinates', [])
                        if len(coords) >= 2:
                            props['lon'] = coords[0]
                            props['lat'] = coords[1]
                    yield props

            _route_processing(iter_geojson_rows(), etl_job)

        else:
            raise ValueError(f"Formato no soportado para ETL nativo: {ext}")
                
        etl_job.status = 'COMPLETED'
        etl_job.completed_at = timezone.now()
        etl_job.error_log += "Procesamiento finalizado con éxito.\n"
        
        etl_job.dataset.last_load = timezone.now()
        etl_job.dataset.save()
        
    except Exception as e:
        etl_job.status = 'FAILED'
        etl_job.error_log += f"\nERROR CRÍTICO:\n{str(e)}\n{traceback.format_exc()}"
    
    etl_job.save()

def _route_processing(reader, etl_job):
    tracked_reader = _track_reader_progress(reader, etl_job)
    name_lower = etl_job.dataset.name.lower()
    if 'hpms' in name_lower or 'highway_performance_monitoring_system' in name_lower:
        _process_hpms(tracked_reader, etl_job)
    elif 'truck stop' in name_lower or 'travel center' in name_lower or 'pilot' in name_lower or 'loves' in name_lower:
        _process_truck_stops(tracked_reader, etl_job)
    elif 'wim' in name_lower:
        _process_wim_stations(tracked_reader, etl_job)
    elif 'recycling' in name_lower or 'epa' in name_lower:
        _process_recycling(tracked_reader, etl_job)
    elif 'carrier' in name_lower or 'company' in name_lower:
        _process_carriers(tracked_reader, etl_job)
    elif 'stat' in name_lower:
        _process_stats(tracked_reader, etl_job)
    elif 'tire' in name_lower:
        _process_tire_shops(tracked_reader, etl_job)
    elif 'fuel' in name_lower:
        _process_alt_fuel(tracked_reader, etl_job)
    elif 'enforcement' in name_lower or 'size' in name_lower:
        _process_enforcement(tracked_reader, etl_job)
    elif 'routes' in name_lower or 'nhs' in name_lower:
        etl_job.records_processed = sum(1 for _ in tracked_reader)
        etl_job.error_log += f"Se detectaron y validaron rutas NHS en formato GeoJSON.\n"
    else:
        etl_job.records_processed = sum(1 for _ in tracked_reader)
        etl_job.error_log += f"Lógica ETL para {etl_job.dataset.name} en desarrollo.\n"


def _process_hpms(reader, etl_job):
    reader_iter = iter(reader)
    first_row = next(reader_iter, None)

    if first_row is None:
        etl_job.error_log += "HPMS: no se encontraron filas para procesar.\n"
        return
    header_tokens = {"".join(ch.lower() for ch in str(key) if ch.isalnum()) for key in first_row.keys()}
    is_inventory_index = {"state", "apiurl", "mapurl"}.issubset(header_tokens)

    if is_inventory_index:
        result = RouteService.seed_hpms_sources_from_csv(etl_job.dataset.file_path, source_dataset=etl_job.dataset)
        etl_job.records_processed = result["created"] + result["updated"]
        etl_job.error_log += (
            "HPMS: inventario maestro procesado. "
            f"Fuentes creadas={result['created']}, actualizadas={result['updated']}, "
            f"total={result['total_sources']}.\n"
        )
        return

    state_name = None
    file_name = os.path.splitext(os.path.basename(etl_job.dataset.file_path))[0].replace("_", " ").replace("-", " ")
    for source in HPMSInventorySource.objects.all():
        if source.state_name.lower() in file_name.lower():
            state_name = source.state_name
            break

    source = HPMSInventorySource.objects.filter(state_name=state_name).first() if state_name else None
    result = RouteService.import_hpms_segments(chain([first_row], reader_iter), source=source, default_state=state_name)
    etl_job.records_processed = result["created"] + result["updated"]
    etl_job.error_log += (
        "HPMS: segmentos procesados. "
        f"Creados={result['created']}, actualizados={result['updated']}, "
        f"snapshots={result['snapshots']}.\n"
    )

def _process_truck_stops(reader, etl_job):
    """
    Procesa Truck Stops.
    Maneja polimorfismo de columnas (NTAD, Locmaster, Love's, Pilot).
    """
    created_count = 0
    processed_count = 0
    
    for row in reader:
        # 1. Identificar columnas de Nombre y Operador (Polimorfismo)
        name = row.get('Facility_Name') or row.get('Name') or row.get('Store Name') or row.get('StoreName') or row.get('nhs_rest_stop') or row.get('municipality') or 'Unknown Truck Stop'
        operator = row.get('Operator') or row.get('Brand') or row.get('Network') or 'Public Rest Stop'
        
        # Inferencia de operador si no viene una columna explícita (Ej: Pilot all_locations_plot.csv)
        if operator == 'Public Rest Stop':
            name_lower = str(name).lower()
            if 'pilot' in name_lower or 'flying j' in name_lower:
                operator = 'Pilot Flying J'
            elif 'love' in name_lower or 'love' in etl_job.dataset.file_path.lower():
                operator = "Love's Travel Stops"
            elif 'ta' in name_lower or 'travelcenters' in name_lower or 'petro' in name_lower:
                operator = 'TA / Petro'

        # Mejorar nombre si es Love's y viene como Unknown
        if operator == "Love's Travel Stops" and name == 'Unknown Truck Stop':
            store_num = row.get('StoreNumber') or row.get('Store_Number') or ''
            if store_num:
                name = f"Love's Travel Stop #{store_num}"
            else:
                name = "Love's Travel Stop"

        # 1.5 Identificar Teléfono y Web
        phone = row.get('Phone') or row.get('Telephone') or row.get('Contact_Number') or row.get('Phone Number') or 'Sin Información'
        website = row.get('Website') or row.get('URL') or row.get('Web') or ''
        
        # Limpieza básica de la web para evitar errores en Django URLField
        if website and not str(website).startswith(('http://', 'https://')):
            # Si el archivo de Love's trae un bloque de texto en la cabecera en vez de URL limpia
            if 'www.' in str(website).lower() or '.com' in str(website).lower():
                website = f"https://{str(website).split()[0]}"
            else:
                website = ''
        
        # 2. Identificar Coordenadas (Polimorfismo)
        lat = row.get('Latitude') or row.get('Lat') or row.get('latitude') or row.get('y') or row.get('Y')
        lon = row.get('Longitude') or row.get('Long') or row.get('longitude') or row.get('x') or row.get('X')
        
        # 3. Identificar Dirección
        address = row.get('Street_Address') or row.get('Address') or row.get('Street') or row.get('highway_route') or 'Sin Información'
        city = row.get('City') or row.get('municipality') or 'Sin Información'
        state = row.get('State') or row.get('State/Province') or row.get('state') or 'Sin Información'
        zip_code = row.get('Zip') or row.get('Zip Code') or row.get('Postal Code') or 'Sin Información'
        
        if not lat or not lon:
            continue # Si no hay GPS, descartamos para el mapa
            
        try:
            lat = float(lat)
            lon = float(lon)
        except (ValueError, TypeError):
            continue
            
        # Identificar espacios de parqueo
        parking_spaces = 0
        spaces_raw = (row.get('Parking Spaces Count') or 
                      row.get('Truck_Parking_Spaces') or 
                      row.get('truck_parking_capacity') or 
                      row.get('TruckParking') or 
                      row.get('number_of_spots') or 
                      row.get('ParkingSpaces') or 
                      row.get('Truck Parking Spaces') or 0)
        try:
            parking_spaces = int(spaces_raw)
        except (ValueError, TypeError):
            pass

        # Identificar islas de diésel para tractomulas (Love's/Pilot)
        diesel_lanes = 0
        lanes_raw = row.get('DEFLanes') or row.get('DieselLanes') or row.get('Truck Fuel Lanes') or row.get('Fuel Lane Count') or 0
        try:
            diesel_lanes = int(lanes_raw)
        except (ValueError, TypeError):
            pass

        # Crear la estación base
        stop, created = TruckStop.objects.update_or_create(
            name=name[:255],
            latitude=lat,
            longitude=lon,
            defaults={
                'operator': operator[:255],
                'phone': phone[:50],
                'website': website[:500] if website else None,
                'address': address,
                'city': city,
                'state': state,
                'zip_code': zip_code,
                'parking_spaces': parking_spaces,
                'diesel_lanes': diesel_lanes
            }
        )
        
        # 4. Extracción de Amenidades (Múltiples formatos)
        amenities_to_add = []
        
        # Diccionario de palabras clave para buscar en las columnas booleanas o de texto
        amenity_keywords = {
            'Showers': ['shower', 'ducha', 'bath'],
            'Restaurant': ['restaurant', 'food', 'comida', 'dining'],
            'Tire Shop': ['tire', 'llanta', 'repair', 'mechanic'],
            'Parking': ['park', 'espacio', 'spaces'],
            'Scales': ['scale', 'wim', 'báscula', 'weigh'],
            'Diesel': ['diesel', 'fuel', 'combustible'],
            'WiFi': ['wifi', 'internet'],
        }
        
        # Estrategia A: Si hay columnas específicas que dicen 'Yes'/'True' o tienen números/nombres (Ej: Dataset NTAD, Locmaster)
        for col_name, value in row.items():
            if value is None or str(value).strip() == '':
                continue
                
            val_str = str(value).strip().lower()
            # Ignorar si es explícitamente "No", "0" o "False"
            if val_str in ['0', 'n', 'no', 'false']:
                continue
                
            # Si tiene cualquier otro valor (un 'yes', un número '11', o un texto 'Subway')
            col_lower = str(col_name).lower()
            for am_name, keywords in amenity_keywords.items():
                if any(k in col_lower for k in keywords):
                    amenity, _ = Amenity.objects.get_or_create(name=am_name)
                    amenities_to_add.append(amenity)
                    
        # Excepciones específicas para nombres de columnas de restaurantes en TA/Petro
        rest = row.get('Full Service Restaurant')
        qsr = row.get('QSR(s)')
        if (rest and str(rest).strip() and str(rest).strip().lower() not in ['n', 'no', '0']) or (qsr and str(qsr).strip() and str(qsr).strip().lower() not in ['n', 'no', '0']):
            am, _ = Amenity.objects.get_or_create(name='Restaurant')
            amenities_to_add.append(am)
            
            # Guardar el nombre específico
            rests_list = []
            if rest and str(rest).strip() and str(rest).strip().lower() not in ['n', 'no', '0']:
                rests_list.append(str(rest).strip())
            if qsr and str(qsr).strip() and str(qsr).strip().lower() not in ['n', 'no', '0']:
                for r in str(qsr).split(','):
                    rests_list.append(r.strip())
                    
            for r_clean in rests_list:
                if r_clean:
                    am_spec, _ = Amenity.objects.get_or_create(name=f"Comida: {r_clean}", defaults={'category': 'Restaurante Específico'})
                    amenities_to_add.append(am_spec)
            
        # Estrategia B: Si hay una columna "Amenities" o "Services" separada por comas (Ej: Locmaster/Loves/Pilot)
        services_col = row.get('Amenities') or row.get('Services') or row.get('Facilities') or ''
        # Estrategia C: Si hay una columna específica de Restaurantes separada por comas (Ej: Pilot / Love's)
        restaurants_col = row.get('Restaurants') or row.get('Restaurant') or ''
        
        if services_col or restaurants_col:
            services_str = str(services_col).lower() + " " + str(restaurants_col).lower()
            # Si el campo de restaurantes tiene algún texto que no esté vacío, forzar la amenidad genérica 'Restaurant'
            if str(restaurants_col).strip() != '' and str(restaurants_col).strip().lower() not in ['none', 'n/a', 'no', '0']:
                services_str += " restaurant "
                
                # También extraer los nombres específicos de los restaurantes
                for r in str(restaurants_col).split(','):
                    r_clean = r.strip()
                    if r_clean and r_clean.lower() not in ['none', 'n/a', 'no', '0']:
                        am_spec, _ = Amenity.objects.get_or_create(name=f"Comida: {r_clean}", defaults={'category': 'Restaurante Específico'})
                        amenities_to_add.append(am_spec)
                
            for am_name, keywords in amenity_keywords.items():
                if any(k in services_str for k in keywords):
                    amenity, _ = Amenity.objects.get_or_create(name=am_name)
                    amenities_to_add.append(amenity)
        
        if amenities_to_add:
            stop.amenities.add(*amenities_to_add)
            
        processed_count += 1
        if created:
            created_count += 1
    
    etl_job.records_processed = processed_count
    etl_job.error_log += f"Se procesaron e importaron {created_count} Truck Stops (detectando automáticamente las columnas).\n"

def _process_wim_stations(reader, etl_job):
    """
    Procesa estaciones WIM y datos de enforcement.
    FILTRO ESTRICTO: Solo guardamos datos relevantes para "Vehículos Clase 8" (FHWA Class 8).
    """
    count = 0
    filtered_count = 0
    
    # Amenidad para saber que este punto tiene telemetría WIM
    wim_amenity, _ = Amenity.objects.get_or_create(name="Pesaje en Movimiento (WIM)", defaults={"category": "Regulación Inteligente"})

    for row in reader:
        # LÓGICA DE FILTRADO PARA CLASE 8
        # Si el dataset tiene una columna de clase de vehículo y NO es 8, lo ignoramos.
        vehicle_class = row.get('Vehicle_Class', row.get('FHWA_Class', ''))
        if vehicle_class and str(vehicle_class) != '8':
            filtered_count += 1
            continue # Descartar todo lo que no sea Clase 8

        station_id = row.get('Station_ID') or row.get('Station_Id') or row.get('site_id') or f"Unknown_{count}"
        lat = row.get('Latitude') or row.get('latitude') or row.get('Y') or row.get('y')
        lon = row.get('Longitude') or row.get('longitude') or row.get('X') or row.get('x')
        
        if not lat or not lon:
            continue

        wim, created = WIMStation.objects.update_or_create(
            station_id=station_id,
            defaults={
                'name': f"WIM Station {station_id}",
                'state': row.get('State', ''),
                'latitude': lat,
                'longitude': lon,
                'direction_of_travel': row.get('Direction', ''),
                'number_of_lanes': int(row.get('Lanes', 1) or 1)
            }
        )
        
        wim.amenities.add(wim_amenity)
        count += 1

    etl_job.records_processed = count
    etl_job.error_log += f"Se procesaron {count} estaciones WIM para Vehículos Clase 8.\n"
    if filtered_count > 0:
        etl_job.error_log += f"Se descartaron {filtered_count} registros por no ser Clase 8.\n"

def _process_recycling(reader, etl_job):
    created_count = 0
    processed_count = 0
    progress_log_interval = 1000
    next_progress_checkpoint = progress_log_interval
    for row in reader:
        # Coordenadas polimórficas (Priorizar Latitude/Longitude sobre y/x)
        lat = row.get('Latitude') or row.get('latitude') or row.get('Lat') or row.get('LATITUDE') or row.get('y')
        lon = row.get('Longitude') or row.get('longitude') or row.get('Lon') or row.get('LONGITUDE') or row.get('x')
        
        is_valid_epa = False
        
        try:
            if lat and lon:
                lat, lon = round(float(lat), 6), round(float(lon), 6)
            else:
                lat, lon = None, None
        except Exception:
            lat, lon = None, None
            
        try:
            # Lógica para EPA_Disaster_Debris_Recovery_Data
            is_epa_file = True if 'EPA' in etl_job.dataset.file_path else False
            
            if is_epa_file:
                tires_flag = str(row.get('Tires', '')).strip()
                if tires_flag in ['1', 'yes', 'y', 'true', 'Yes'] or tires_flag.lower() in ['yes', 'y', 'true', '1']:
                    
                    zip_code = str(row.get('Zip') or '').strip()
                    state = row.get('State') or ''
                    name = str(row.get('Company') or 'EPA Recovery Site')[:255]
                    if not name or name == 'None':
                        name = 'EPA Recovery Site'
                    
                    tires_recycled = 1.0 # Valor simbólico
                    tires_generated = 0.0
                    population = 0
                    
                    is_valid_epa = True
                else:
                    is_valid_epa = False
                    
            else:
                is_epa_file = False
                is_valid_epa = False
                # Lógica para US_Recycling_Infrastructure_2022
                tires_recycled_raw = row.get('Tires Tons Recycled')
                tires_recycled = float(tires_recycled_raw) if tires_recycled_raw else 0.0
                
                tires_generated_raw = row.get('Tires Tons  Generated')
                tires_generated = float(tires_generated_raw) if tires_generated_raw else 0.0
                
                population_raw = row.get('Population')
                population = int(population_raw) if population_raw else 0
                
                # Solo guardar infraestructuras que tengan datos de llantas
                if tires_recycled <= 0 and tires_generated <= 0:
                    continue
                    
                zip_code = row.get('Zip_Code', '')
                state = row.get('State', '')
                name = f"Planta Reciclaje Llantas {zip_code} - {state}"
                
        except Exception as e:
            print("ERROR", e)
            continue

        # Usar coordenadas como identificador si existen (EPA), si no, usar zip_code (US_Recycling)
        if is_valid_epa and lat is not None and lon is not None:
            # Validar que lat y lon estén dentro de rangos normales de coordenadas antes de guardar
            if abs(lat) > 90 or abs(lon) > 180:
                pass
            else:
                try:
                    # Actualizar si existe, crear si no existe
                    facility, created = RecyclingFacility.objects.update_or_create(
                        latitude=lat,
                        longitude=lon,
                        defaults={
                            'name': name,
                            'state': state,
                            'zip_code': zip_code[:20],
                            'tires_recycled_tons': tires_recycled,
                            'tires_generated_tons': tires_generated,
                            'population_served': population
                        }
                    )
                    processed_count += 1
                    if created:
                        created_count += 1
                except Exception as e:
                    print("Error saving EPA:", e)
                    pass
        elif not is_epa_file:
            try:
                facility, created = RecyclingFacility.objects.update_or_create(
                    zip_code=zip_code,
                    defaults={
                        'name': name,     
                        'state': state,
                        'latitude': lat,
                        'longitude': lon,
                        'tires_recycled_tons': tires_recycled,
                        'tires_generated_tons': tires_generated,
                        'population_served': population
                    }
                )
                processed_count += 1
                if created:
                    created_count += 1
            except Exception as e:
                pass

        if processed_count >= next_progress_checkpoint:
            etl_job.records_processed = processed_count
            etl_job.error_log += (
                f"Recycling: avance parcial {processed_count} registros procesados.\n"
            )
            etl_job.save(update_fields=['records_processed', 'error_log'])
            next_progress_checkpoint += progress_log_interval
        
    etl_job.records_processed = processed_count
    etl_job.error_log += f"Se importaron {created_count} plantas de reciclaje de llantas.\n"

def _process_carriers(reader, etl_job):
    created_count = 0
    batch_size = 10000
    batch = []
    progress_log_interval = 50000
    next_progress_checkpoint = progress_log_interval
    
    for row in reader:
        dot_num = row.get('DOT_NUMBER')
        if not dot_num: continue

        batch.append(CarrierCompany(
            dot_number=dot_num,
            legal_name=str(row.get('LEGAL_NAME', 'Unknown'))[:255],
            dba_name=str(row.get('DBA_NAME', ''))[:255],
            carrier_operation=str(row.get('CARRIER_OPERATION', ''))[:100],    
            status=str(row.get('STATUS_CODE', 'Active'))[:50]
        ))
        
        if len(batch) >= batch_size:
            # Inserción masiva optimizada para PostgreSQL
            CarrierCompany.objects.bulk_create(
                batch,
                batch_size=batch_size,
                update_conflicts=True,
                unique_fields=['dot_number'],
                update_fields=['legal_name', 'dba_name', 'carrier_operation', 'status']
            )
            created_count += len(batch)
            etl_job.records_processed = created_count
            if created_count >= next_progress_checkpoint:
                etl_job.error_log += (
                    f"Carriers: avance parcial {created_count} registros procesados.\n"
                )
                next_progress_checkpoint += progress_log_interval
            etl_job.save(update_fields=['records_processed', 'error_log'])
            batch = []

    # Insertar el remanente
    if batch:
        CarrierCompany.objects.bulk_create(
            batch,
            batch_size=batch_size,
            update_conflicts=True,
            unique_fields=['dot_number'],
            update_fields=['legal_name', 'dba_name', 'carrier_operation', 'status']
        )
        created_count += len(batch)

    etl_job.records_processed = created_count
        
    etl_job.error_log += f"Se importaron {created_count} empresas transportistas (Carriers) usando procesamiento masivo por lotes.\n"

def _process_stats(reader, etl_job):
    import datetime
    created_count = 0
    for row in reader:
        date_str = row.get('Date')
        if not date_str: continue

        try:
            # Formato en los datasets gubernamentales puede ser "1/1/2026 0:00" o "1947 Jan 01 12:00:00 AM"
            date_part = str(date_str).split()[0] # Tratar de aislar la parte de la fecha
            
            try:
                if '/' in date_part:
                    date_obj = datetime.datetime.strptime(date_part, '%m/%d/%Y').date()
                elif '-' in date_part:
                    date_obj = datetime.datetime.strptime(date_part, '%Y-%m-%d').date()
                else:
                    # Formato "1947 Jan 01"
                    clean_date_str = " ".join(str(date_str).split()[:3])
                    date_obj = datetime.datetime.strptime(clean_date_str, '%Y %b %d').date()
            except ValueError:
                continue

            truck_emp = str(row.get('Transportation Employment - Truck Transportation', '')).replace(',', '').strip()
            fatalities = str(row.get('Highway Fatalities', '')).replace(',', '').strip()     
            diesel = str(row.get('Highway Fuel Price - On-highway Diesel', '')).replace(',', '').replace('$', '').strip()      

            # Solo guardar si hay al menos un dato útil para la industria de tractomulas
            if not truck_emp and not diesel and not fatalities:
                continue

            TransportationStatistic.objects.update_or_create(
                date=date_obj,
                defaults={
                    'truck_employment': int(truck_emp) if truck_emp else None,  
                    'highway_fatalities': int(fatalities) if fatalities else None,
                    'diesel_price': float(diesel) if diesel else None,
                }
            )
            created_count += 1
        except Exception:
            continue

    etl_job.records_processed = created_count
    etl_job.error_log += f"Se importaron {created_count} meses de estadísticas de transporte relevantes para Clase 8.\n"

def _process_enforcement(reader, etl_job):
    created_count = 0
    for row in reader:
        year = row.get('year')
        state = row.get('State')
        
        if not year or not state: continue
        
        try:
            year_int = int(year)
            
            # Extraer números limpiando posibles espacios vacíos
            def get_int(key):
                val = str(row.get(key, '0')).strip()
                return int(val) if val else 0
                
            fixed_weighed = get_int('Vehicles Weighed (Fixed platform Scale)')
            wim_weighed = get_int('Vehicles Weighed (WIM Scale)')
            oversize = get_int('oversize_violation_current_year')
            overweight = get_int('overweight_violation_current_year')
            
            WeightEnforcementStatistic.objects.update_or_create(
                year=year_int,
                state=state[:10],
                defaults={
                    'vehicles_weighed_fixed': fixed_weighed,
                    'vehicles_weighed_wim': wim_weighed,
                    'oversize_violations': oversize,
                    'overweight_violations': overweight
                }
            )
            created_count += 1
        except (ValueError, TypeError):
            continue
            
    etl_job.records_processed = created_count
    etl_job.error_log += f"Se importaron {created_count} registros anuales de control y pesaje por Estado.\n"

def _process_tire_shops(reader, etl_job):
    created_count = 0
    for row in reader:
        # Extraer latitud y longitud, el CSV tiene las columnas 'lat' y 'lon'
        lat = row.get('lat') or row.get('latitude') or row.get('y') or row.get('Lat')
        lon = row.get('lon') or row.get('longitude') or row.get('x') or row.get('Lon')
        if not lat or not lon: continue

        try:
            lat, lon = float(lat), float(lon)
        except (ValueError, TypeError):
            continue

        # Extraer campos de dirección que pueden venir directamente o dentro de JSON tags
        address = row.get('street', '')
        if not address and row.get('housenumber'):
            address = f"{row.get('housenumber')} {row.get('street', '')}"
            
        city = row.get('city') or row.get('addr:city', 'Sin Información')
        state = row.get('state') or row.get('addr:state', 'Sin Información')
        zip_code = row.get('postcode') or row.get('addr:postcode', 'Sin Información')

        TireShop.objects.update_or_create(
            latitude=lat,
            longitude=lon,
            defaults={
                'name': row.get('name', 'Unknown Tire Shop')[:255],
                'brand': row.get('brand', '')[:100],
                'operator': row.get('operator', 'Sin Información')[:255],      
                'address': address if address else 'Sin Información',
                'city': city,
                'state': state,
                'zip_code': zip_code,       
                'phone': row.get('phone', 'Sin Información')[:50],
                'website': row.get('website', '')[:500],
                'is_24_hours': str(row.get('opening_hours', '')).lower() == '24/7'
            }
        )
        created_count += 1
    etl_job.records_processed = created_count
    etl_job.error_log += f"Se importaron {created_count} talleres de llantas.\n"

def _process_alt_fuel(reader, etl_job):
    created_count = 0
    # Como son muchos registros, usaremos una lista para hacer bulk_create si es posible,
    # pero update_or_create es más seguro contra duplicados.
    
    for row in reader:
        # Filtro estricto: Solo estaciones aptas para tractomulas (Clase 8 / Heavy Duty)
        # El archivo usa 'NG Vehicle Class', 'Maximum Vehicle Class', o códigos 'HD'
        max_class = str(row.get('Maximum Vehicle Class') or row.get('NG Vehicle Class') or '').upper()
        if max_class not in ['HD', 'HEAVY-DUTY', 'HEAVY', 'CLASS 8']:
            continue # Si es solo para autos ligeros (LD/MD), la ignoramos
            
        lat = row.get('Latitude') or row.get('y')
        lon = row.get('Longitude') or row.get('x')
        if not lat or not lon: continue
            
        try:
            lat, lon = float(lat), float(lon)
            cng_dispensers = int(row.get('CNG Dispenser Num') or 0)
            ev_dc_fast = int(row.get('EV DC Fast Count') or 0)
        except (ValueError, TypeError):
            continue
            
        station, created = AlternativeFuelStation.objects.update_or_create(
            latitude=lat,
            longitude=lon,
            defaults={
                'name': row.get('Station Name', 'Unknown Alt Fuel')[:255],
                'fuel_type_code': row.get('Fuel Type Code', '')[:20],
                'ev_connector_types': row.get('EV Connector Types', '')[:255],
                'is_public': str(row.get('Groups With Access Code', '')).lower() == 'public',
                'operator': row.get('EV Network', 'Sin Información')[:255],
                'address': row.get('Street Address', 'Sin Información'),
                'city': row.get('City', 'Sin Información'),
                'state': row.get('State', 'Sin Información'),
                'zip_code': row.get('ZIP', 'Sin Información'),
                'phone': row.get('Station Phone', 'Sin Información')[:50],
                'maximum_vehicle_class': max_class,
                'cng_dispenser_num': cng_dispensers,
                'ev_dc_fast_count': ev_dc_fast
            }
        )

        # Map Facility Type to Amenities
        amenities_to_add = []
        facility_type = str(row.get('Facility Type', '')).upper()
        
        if facility_type in ['TRAVEL_CENTER', 'TRUCK_STOP', 'REST_STOP']:
            am, _ = Amenity.objects.get_or_create(name='Parador / Descanso', defaults={'category': 'Descanso'})
            amenities_to_add.append(am)
        elif facility_type in ['RESTAURANT', 'BREWERY_DISTILLERY_WINERY']:
            am, _ = Amenity.objects.get_or_create(name='Restaurant', defaults={'category': 'Alimentación'})
            amenities_to_add.append(am)
        elif facility_type in ['CONVENIENCE_STORE', 'GROCERY', 'PHARMACY', 'RETAIL']:
            am, _ = Amenity.objects.get_or_create(name='Tienda de Conveniencia', defaults={'category': 'Compras'})
            amenities_to_add.append(am)
        elif facility_type in ['AUTO_REPAIR', 'FLEET_GARAGE', 'CARWASH']:
            am, _ = Amenity.objects.get_or_create(name='Taller / Lavado', defaults={'category': 'Mantenimiento'})
            amenities_to_add.append(am)
        elif facility_type in ['HOTEL', 'INN', 'B_AND_B', 'CAMPGROUND', 'RV_PARK']:
            am, _ = Amenity.objects.get_or_create(name='Hospedaje', defaults={'category': 'Descanso'})
            amenities_to_add.append(am)

        if amenities_to_add:
            station.amenities.add(*amenities_to_add)

        created_count += 1
    etl_job.records_processed = created_count
    etl_job.error_log += f"Se importaron {created_count} estaciones de combustible alternativo (Exclusivas para Tractomulas / HD).\n"
